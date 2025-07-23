import streamlit as st
import pandas as pd
from datetime import datetime
import requests
from openpyxl import load_workbook
import io
from msal import PublicClientApplication

# === KONFIGURASJON ===
CLIENT_ID = "Da9606e82-6f2d-432a-84f5-f3ce46e65413"
TENANT_ID = "6be85a4e-06c8-4985-94a3-074807b81b19"
AUTHORITY = f"https://login.microsoftonline.com/{TENANT_ID}"
SCOPES = ["Files.ReadWrite.All", "offline_access", "User.Read"]
EXCEL_FILENAME = "vedlikeholdsplan_ver22.xlsx"
EXCEL_SHEET = "Erfaringslogg"

# === AUTENTISERING ===
def get_access_token():
    app = PublicClientApplication(CLIENT_ID, authority=AUTHORITY)
    accounts = app.get_accounts()
    if accounts:
        result = app.acquire_token_silent(SCOPES, account=accounts[0])
    else:
        flow = app.initiate_device_flow(scopes=SCOPES)
        if "user_code" not in flow:
            raise Exception("Kan ikke starte enhetsflyt")
        st.info(f"Gå til {flow['verification_uri']} og skriv inn koden: {flow['user_code']}")
        result = app.acquire_token_by_device_flow(flow)
    return result["access_token"]

# === LAST NED OG LAST OPP EXCEL ===
def download_excel(access_token):
    headers = {"Authorization": f"Bearer {access_token}"}
    url = f"https://graph.microsoft.com/v1.0/me/drive/root:/Documents/{EXCEL_FILENAME}:/content"
    response = requests.get(url, headers=headers)
    return io.BytesIO(response.content)

def upload_excel(access_token, file_bytes):
    headers = {
        "Authorization": f"Bearer {access_token}",
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    }
    url = f"https://graph.microsoft.com/v1.0/me/drive/root:/Documents/{EXCEL_FILENAME}:/content"
    response = requests.put(url, headers=headers, data=file_bytes.getvalue())
    return response.status_code == 200

# === HENT VÆRDATA ===
def hent_vaerdata(lat, lon):
    try:
        url = (
            f"https://api.open-meteo.com/v1/forecast?"
            f"latitude={lat}&longitude={lon}&current=temperature_2m,weathercode,windspeed_10m"
            f"&timezone=auto"
        )
        response = requests.get(url)
        data = response.json()
        temp = data["current"]["temperature_2m"]
        wind = data["current"]["windspeed_10m"]
        weather_code = data["current"]["weathercode"]
        weather_map = {
            0: "Klar himmel", 1: "Hovedsakelig klar", 2: "Delvis skyet", 3: "Overskyet",
            45: "Tåke", 48: "Tåke med rim", 51: "Lett yr", 53: "Moderat yr", 55: "Kraftig yr",
            61: "Lett regn", 63: "Moderat regn", 65: "Kraftig regn", 71: "Lett snø",
            73: "Moderat snø", 75: "Kraftig snø"
        }
        weather = weather_map.get(weather_code, "Ukjent vær")
        return weather, temp, f"{wind} m/s"
    except:
        return "Ukjent", "Ukjent", "Ukjent"

# === APP ===
st.set_page_config(page_title="Erfaringslogg", layout="centered")
st.title("📋 Registrer tiltak og erfaring")

today = datetime.today().strftime("%Y-%m-%d")
latitude = 70.1112
longitude = 29.3532
vaer, temperatur, vind = hent_vaerdata(latitude, longitude)

try:
    token = get_access_token()
    excel_io = download_excel(token)
    df_plan = pd.read_excel(excel_io, sheet_name="Vedlikeholdsplan", engine="openpyxl")
    tiltak_liste = df_plan["Tiltak"].dropna().unique().tolist()
except Exception as e:
    st.error("Feil ved lasting av Excel-fil fra OneDrive.")
    st.stop()

with st.form("erfaringsskjema"):
    st.markdown(f"**📅 Dato:** {today}")
    st.markdown(f"**🌤️ Vær:** {vaer}")
    st.markdown(f"**🌡️ Temperatur:** {temperatur}")
    st.markdown(f"**💨 Vind:** {vind}")

    tiltak = st.selectbox("Tiltak", tiltak_liste)
    utført_av = st.text_input("Utført av")
    timer = st.number_input("Timer brukt", min_value=0.0, step=0.5)
    erfaring = st.text_area("Erfaring")
    forbedringer = st.text_area("Forslag til forbedringer")

    send = st.form_submit_button("📤 Lagre")

if send:
    try:
        excel_io.seek(0)
        wb = load_workbook(excel_io)
        sheet = wb[EXCEL_SHEET]

        row = 2
        while sheet.cell(row=row, column=1).value not in [None, ""]:
            row += 1

        sheet.cell(row=row, column=1).value = today
        sheet.cell(row=row, column=2).value = vaer
        sheet.cell(row=row, column=3).value = temperatur
        sheet.cell(row=row, column=4).value = vind
        sheet.cell(row=row, column=5).value = tiltak
        sheet.cell(row=row, column=6).value = utført_av
        sheet.cell(row=row, column=7).value = timer
        sheet.cell(row=row, column=8).value = erfaring
        sheet.cell(row=row, column=9).value = forbedringer

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        if upload_excel(token, output):
            st.success("✅ Registreringen er lagret i OneDrive.")
        else:
            st.error("❌ Kunne ikke laste opp til OneDrive.")
    except Exception as e:
        st.error("Feil ved lagring av data.")
