import streamlit as st
import pandas as pd
from datetime import datetime
import requests
from openpyxl import load_workbook
import io

# === KONFIGURASJON ===
EXCEL_FILE_URL = "https://arcticstorage-my.sharepoint.com/personal/bjornarne_arcticstorage_no/Documents/vedlikeholdsplan_ver22%201.xlsx"
EXCEL_SHEET = "Erfaringslogg"

# === HENT VÆRDATA ===
def hent_vaerdata(lat, lon):
    try:
        url = (
            f"https://api.open-meteo.com/v1/forecast?"
            f"latitude={lat}&longitude={lon}&current=temperature_2m,weathercode,windspeed_10m"
            f"&timezone=auto"
        )
        response = requests.get(url)
        data = response.json()
        temp = data["current"]["temperature_2m"]
        wind = data["current"]["windspeed_10m"]
        weather_code = data["current"]["weathercode"]
        weather_map = {
            0: "Klar himmel", 1: "Hovedsakelig klar", 2: "Delvis skyet", 3: "Overskyet",
            45: "Tåke", 48: "Tåke med rim", 51: "Lett yr", 53: "Moderat yr", 55: "Kraftig yr",
            61: "Lett regn", 63: "Moderat regn", 65: "Kraftig regn", 71: "Lett snø",
            73: "Moderat snø", 75: "Kraftig snø"
        }
        weather = weather_map.get(weather_code, "Ukjent vær")
        return weather, temp, f"{wind} m/s"
    except:
        return "Ukjent", "Ukjent", "Ukjent"

# === LAST NED EXCEL-FIL FRA SHAREPOINT ===
def download_excel_from_sharepoint(url):
    try:
        response = requests.get(url)
        if response.status_code == 200:
            return io.BytesIO(response.content)
        else:
            st.error(f"Kunne ikke hente Excel-fil. Statuskode: {response.status_code}")
            return None
    except Exception as e:
        st.error(f"Feil ved nedlasting av Excel-fil: {e}")
        return None

# === APP ===
st.set_page_config(page_title="Erfaringslogg", layout="centered")
st.title("📋 Registrer tiltak og erfaring")

today = datetime.today().strftime("%Y-%m-%d")
latitude = 70.1112
longitude = 29.3532
vaer, temperatur, vind = hent_vaerdata(latitude, longitude)

excel_io = download_excel_from_sharepoint(EXCEL_FILE_URL)
if excel_io is None:
    st.stop()

try:
    df_plan = pd.read_excel(excel_io, sheet_name="Vedlikeholdsplan", engine="openpyxl")
    tiltak_liste = df_plan["Tiltak"].dropna().unique().tolist()
except Exception as e:
    st.error("Feil ved lesing av vedlikeholdsplan fra Excel.")
    st.stop()

with st.form("erfaringsskjema"):
    st.markdown(f"**📅 Dato:** {today}")
    st.markdown(f"**🌤️ Vær:** {vaer}")
    st.markdown(f"**🌡️ Temperatur:** {temperatur}")
    st.markdown(f"**💨 Vind:** {vind}")

    tiltak = st.selectbox("Tiltak", tiltak_liste)
    utført_av = st.text_input("Utført av")
    timer = st.number_input("Timer brukt", min_value=0.0, step=0.5)
    erfaring = st.text_area("Erfaring")
    forbedringer = st.text_area("Forslag til forbedringer")

    send = st.form_submit_button("📤 Lagre")

if send:
    try:
        excel_io.seek(0)
        wb = load_workbook(excel_io)
        sheet = wb[EXCEL_SHEET]

        row = 2
        while sheet.cell(row=row, column=1).value not in [None, ""]:
            row += 1

        sheet.cell(row=row, column=1).value = today
        sheet.cell(row=row, column=2).value = vaer
        sheet.cell(row=row, column=3).value = temperatur
        sheet.cell(row=row, column=4).value = vind
        sheet.cell(row=row, column=5).value = tiltak
        sheet.cell(row=row, column=6).value = utført_av
        sheet.cell(row=row, column=7).value = timer
        sheet.cell(row=row, column=8).value = erfaring
        sheet.cell(row=row, column=9).value = forbedringer

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        st.success("✅ Registreringen er lagret lokalt. (OBS: Opplasting til SharePoint krever egen løsning)")
    except Exception as e:
        st.error(f"Feil ved lagring av data: {e}")
